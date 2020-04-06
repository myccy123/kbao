from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import NamedStyle


def read_excel(file_path, sheet=None, read_only=True, limit=0,
               skiprow=0, skipcol=0):
    wb = load_workbook(file_path, read_only=read_only)
    if sheet is None:
        ws = wb.active
    else:
        ws = wb[sheet]
    res_data = []
    for i, row in enumerate(ws.iter_rows()):
        if i < skiprow:
            continue
        if 0 < limit == i:
            break
        d = []
        for j, cell in enumerate(row):
            if j < skipcol:
                continue
            d.append(cell.value)
        res_data.append(d)
    wb.close()
    return res_data


def read_sheet_list(file_path):
    wb = load_workbook(file_path, read_only=True)
    res = wb.sheetnames
    wb.close()
    return res


def gen_col_name(length):
    j = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
         'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    def loop(s):
        if s == 'Z':
            return 'A'
        else:
            return j[j.index(s) + 1]

    def setp_up(old, idx=0):
        if idx == 0:
            tmp = list(reversed(old))
        else:
            tmp = old

        if tmp[idx] == 'Z':
            tmp[idx] = 'A'
            if idx == len(tmp) - 1:
                tmp.insert(0, 'A')
            else:
                setp_up(tmp, idx + 1)
        else:
            tmp[idx] = loop(tmp[idx])
        return list(reversed(tmp))

    def get_next(s):
        if s == '':
            return 'A'
        else:
            tmp = list(s)
            if tmp[-1] == 'Z':
                return ''.join(setp_up(tmp))
            else:
                tmp[-1] = loop(tmp[-1])
                return ''.join(tmp)

    last = ''
    res = []
    i = 0
    while i < length:
        last = get_next(last)
        res.append(last)
        i += 1

    return res


def write_excel(filename='export.xlsx', rows=None, hearders=None):
    if hearders is None:
        hearders = []
    if rows is None:
        rows = []
    wb = Workbook()
    ws = wb.active
    has_header = True if hearders else False
    has_row = True if rows else False

    if has_header and has_row:
        if len(hearders) != len(rows[0]):
            raise Exception()

    if has_header:
        rows.insert(0, hearders)

    for i, row in enumerate(rows):
        for j, col in enumerate(row):
            ws.cell(row=i + 1, column=j + 1, value=col)

    print(filename, '已生成!')
    wb.save(filename)
